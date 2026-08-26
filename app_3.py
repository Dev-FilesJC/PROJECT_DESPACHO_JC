import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date, datetime
from contextlib import contextmanager
import time
import json
import uuid
from pathlib import Path
import os
import threading
import base64
import requests
from openpyxl import load_workbook

st.set_page_config(
    page_title="JC Control de Solicitudes — Despacho",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# =========================================================
# ESTILOS
# =========================================================
st.markdown(
    """
<style>
* { box-sizing: border-box; }

.stApp {
    background:#020914;
    color:#f3f4f6;
    font-family:Arial,Helvetica,sans-serif;
    font-size:13px;
}
.block-container {
    width:94%;
    max-width:1450px;
    padding-top:22px !important;
    padding-bottom:24px !important;
    margin:auto;
}
header[data-testid="stHeader"] { background:transparent; }
div[data-testid="stToolbar"] { display:none; }
footer { display:none; }

h1,h2,h3 {
    color:#f3f4f6 !important;
    font-family:Arial,Helvetica,sans-serif !important;
}
.titulo {
    text-align:center;
    margin:4px 0 14px;
    font-size:30px;
    font-weight:900;
    line-height:1.2;
    text-align:center;
    width:100%;
    display:block;
    white-space:nowrap;
}
.subtitulo {
    margin:0 0 12px;
    font-size:13px;
    font-weight:700;
}

.stButton > button,
.stDownloadButton > button {
    border:1px solid #273246 !important;
    border-radius:5px !important;
    min-height:26px !important;
    height:26px !important;
    padding:4px 7px !important;
    font-weight:800 !important;
    color:white !important;
    background:#17263d !important;
    font-size:9.5px !important;
}
.stButton > button:hover,
.stDownloadButton > button:hover {
    filter:brightness(1.18);
    border-color:#315b86 !important;
}

div[data-testid="stFileUploader"] { margin-top:-5px; }
div[data-testid="stFileUploader"] section {
    background:#252630 !important;
    border:1px solid #303442 !important;
    border-radius:5px !important;
    padding:4px 8px !important;
    min-height:38px !important;
}
div[data-testid="stFileUploader"] section > div { padding:0 !important; }
div[data-testid="stFileUploader"] button {
    background:#123a2d !important;
    border:1px solid #1b654c !important;
    color:white !important;
    border-radius:5px !important;
    font-size:11px !important;
    font-weight:800 !important;
}

div[data-testid="stForm"] {
    background:rgba(7,15,29,.55);
    border:1px solid #273246;
    border-radius:6px;
    padding:9px !important;
}
div[data-testid="stForm"] label {
    color:#cbd5e1 !important;
    font-size:10px !important;
    font-weight:700 !important;
}
div[data-baseweb="input"],
div[data-baseweb="select"] > div,
div[data-testid="stDateInput"] > div {
    background:#242630 !important;
    color:#f3f4f6 !important;
    border-radius:5px !important;
    border-color:transparent !important;
    min-height:33px !important;
}
div[data-baseweb="input"] input,
div[data-testid="stDateInput"] input {
    color:#f3f4f6 !important;
    font-size:11px !important;
}
div[data-baseweb="select"] span {
    color:#f3f4f6 !important;
    font-size:11px !important;
}
div[data-baseweb="select"] svg { fill:#f3f4f6 !important; }

div[data-testid="stAlert"] {
    border-radius:5px !important;
    font-size:11px !important;
}

.metric-card {
    min-height:46px;
    padding:4px 2px;
    background:transparent;
    border:0;
    text-align:left;
}
.metric-card .label {
    display:block;
    color:#b5bcc8;
    font-size:10px;
}
.metric-card .value {
    font-size:18px;
    font-weight:500;
    color:#f3f4f6;
}

div[data-testid="stTextInput"] input {
    width:100%;
    height:33px;
    border:1px solid transparent;
    border-radius:5px;
    background:#242630 !important;
    color:#f3f4f6 !important;
    padding:0 9px;
    font-size:11px;
}

/* Selector de estado compacto junto al buscador */
div[data-testid="stSelectbox"] > div {
    min-height:33px !important;
}
div[data-testid="stSelectbox"] [data-baseweb="select"] > div {
    min-height:33px !important;
    padding-top:0 !important;
    padding-bottom:0 !important;
}

div[data-testid="stDataFrame"] {
    width:100%;
    border:1px solid #273246 !important;
    border-radius:6px;
    overflow:hidden;
}

hr { border-color:#273246 !important; }

.footer {
    margin-top:30px;
    padding:15px 10px;
    border-top:1px solid #273246;
    text-align:center;
    color:#8f9bad;
    font-size:11px;
}
.footer strong { color:#dbe2ea; }

@media (max-width:800px) {
    .block-container {
        width:94%;
        padding:10px !important;
    }
}
</style>
""",
    unsafe_allow_html=True,
)

COLS = [
    "CLIENTE",
    "NRO SOLICITUD - WO",
    "TIPO DE SOLICITUD",
    "PRIORIDAD",
    "CANT - ITEMS",
    "ESTADO DE SOLICITUD",
    "DIRECCIÓN",
    "FECHA DE INGRESO",
]
RUTA_SHEET = "PROGRAMACION_RUTAS"

CLIENTES = [
    "Seleccione una opcion",
    "BANCO SOL",
    "BANCO BNB",
    "BANCO FIE",
    "BANCO FORTALEZA",
    "PRENDAMAS",
    "MOLINO ANDINO",
    "PREVICOR CORREDORES",
    "WCS-BOLIVIA",
    "INDUSTRIA Y COMERCIO ALICONSUMO",
    "LABORATORIOS LKM",
    "CLINICA LOS OLIVOS",
]
TIPOS = [
    "Seleccione una opcion",
    "EXTERNO",
    "INTERNO - INV.",
    "INTERNO - BPO.",
    "INDEXACION - BPO.",
    "REVISION INTERNA",
    "SERVICIOS",
    "ENVIO DE CAJAS NUEVAS",
]
PRIORIDADES = ["Seleccione una opcion", "RUSH", "TURNO SIGUIENTE"]
ESTADOS = [
    "Seleccione una opcion",
    "POR EXTRAER",
    "POR REASIGNAR",
    "POR ENVIAR",
    "ENTREGADO",
    "ENVIADO",
    "ANULADO",
    "POR ETIQUETAR",
]
DIRECCIONES = [
    "Seleccione una opcion",
    "POLYSISTEMAS",
    "EVARISTO VALLE",
    "LA PAZ - CAMACHO",
    "SAN PEDRO OF NAL",
    "SAN MIGUEL",
    "ZONA SUR",
    "12 DE OCTUBLE",
    "PALENQUE",
    "SATELITE",
    "COCHABAMBA",
    "SANTA CRUZ",
    "SUCRE",
    "ORURO",
    "POTOSI",
    "Banco Sol",
    "Banco BNB",
    "Banco Fie",
    "Molino Andino",
    "Prendamas",
    "Banco Fortaleza",
    "Previcor Corredores",
    "Wcs - Bolivia",
    "Industria y comercio Aliconsumo",
    "Banco Nacional de Bolivia - La Paz",
    "Banco Nacional de Bolivia - Sucre",
    "LABORATORIOS LKM",
    "CLINICA LOS OLIVOS",
]

if "rows" not in st.session_state:
    st.session_state.rows = None
if "rutas" not in st.session_state:
    st.session_state.rutas = None
if "filename" not in st.session_state:
    st.session_state.filename = ""
if "excel_path" not in st.session_state:
    st.session_state.excel_path = ""
if "editing" not in st.session_state:
    st.session_state.editing = None
if "editing_key" not in st.session_state:
    st.session_state.editing_key = ""
if "page" not in st.session_state:
    st.session_state.page = "solicitudes"
if "form_version" not in st.session_state:
    st.session_state.form_version = 0
if "tabla_version" not in st.session_state:
    st.session_state.tabla_version = 0
if "session_id" not in st.session_state:
    st.session_state.session_id = uuid.uuid4().hex
if "registro_activo" not in st.session_state:
    st.session_state.registro_activo = False
if "usuario_nombre" not in st.session_state:
    st.session_state.usuario_nombre = ""
if "uploader_version" not in st.session_state:
    st.session_state.uploader_version = 0
if "logged_out" not in st.session_state:
    st.session_state.logged_out = False
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
if "usuario_login" not in st.session_state:
    st.session_state.usuario_login = ""


@st.cache_resource
def obtener_recurso_cloud():
    return {
        "lock": threading.RLock(),
        "owner_id": None,
        "usuario": None,
        "hora_inicio": None,
        "motivo": None,
        "write_owner": None,
    }


CLOUD = obtener_recurso_cloud()
CLOUD_PATH = Path("/tmp/jc_control_solicitudes_shared.xlsx")

# =========================================================
# GITHUB — EXCEL CENTRALIZADO
# =========================================================
GITHUB_OWNER = "Dev-FilesJC"
GITHUB_REPO = "PROJECT_DESPACHO_JC"
GITHUB_BRANCH = "main"
GITHUB_EXCEL = "CONTROL-SOLICITUDES-ALMACEN.xlsx"
GITHUB_API = "https://api.github.com"


def github_configurado():
    try:
        return bool(st.secrets.get("GITHUB_TOKEN", ""))
    except Exception:
        return False


def github_headers():
    token = st.secrets.get("GITHUB_TOKEN", "")
    if not token:
        raise RuntimeError("Falta GITHUB_TOKEN en los Secrets de Streamlit Cloud.")
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }


def github_url():
    return f"{GITHUB_API}/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{GITHUB_EXCEL}"


def descargar_excel_github(destino=CLOUD_PATH):
    r = requests.get(
        github_url(), headers=github_headers(),
        params={"ref": GITHUB_BRANCH}, timeout=30
    )
    if r.status_code == 404:
        raise FileNotFoundError(
            f"No existe '{GITHUB_EXCEL}' en {GITHUB_OWNER}/{GITHUB_REPO}. "
            "Sube primero el Excel inicial al repositorio."
        )
    r.raise_for_status()
    data = r.json()
    content = data.get("content", "").replace("\n", "")
    if not content:
        raise RuntimeError("GitHub no devolvió el contenido del Excel.")
    Path(destino).write_bytes(base64.b64decode(content))
    return data.get("sha")


def subir_excel_github(ruta, mensaje="Actualizar solicitudes desde Streamlit"):
    ruta = Path(ruta)
    r_get = requests.get(
        github_url(), headers=github_headers(),
        params={"ref": GITHUB_BRANCH}, timeout=30
    )
    if r_get.status_code == 404:
        raise FileNotFoundError(f"No existe '{GITHUB_EXCEL}' en GitHub.")
    r_get.raise_for_status()
    sha = r_get.json().get("sha")
    payload = {
        "message": mensaje,
        "content": base64.b64encode(ruta.read_bytes()).decode("ascii"),
        "branch": GITHUB_BRANCH,
        "sha": sha,
    }
    r_put = requests.put(
        github_url(), headers=github_headers(), json=payload, timeout=60
    )
    if r_put.status_code == 409:
        raise TimeoutError(
            "Otro usuario actualizó el Excel en GitHub al mismo tiempo. "
            "Pulsa ACTUALIZAR PAGINA y vuelve a guardar."
        )
    if r_put.status_code in (401, 403):
        raise PermissionError(
            "GitHub rechazó la escritura. Verifica GITHUB_TOKEN y el permiso "
            "Contents: Read and write."
        )
    r_put.raise_for_status()
    return r_put.json()


def normalizar(v):
    return str(v if pd.notna(v) else "").strip().upper()


def cargar_excel_desde_ruta(ruta):
    """Lee directamente el Excel elegido por el usuario."""
    xls = pd.ExcelFile(ruta)
    sheet = "SOLICITUDES" if "SOLICITUDES" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(ruta, sheet_name=sheet, dtype=str).fillna("")

    rutas = None
    if RUTA_SHEET in xls.sheet_names:
        rutas = pd.read_excel(ruta, sheet_name=RUTA_SHEET, dtype=str).fillna("")

    salida = pd.DataFrame(columns=COLS)
    for c in COLS:
        encontrada = next(
            (x for x in df.columns if normalizar(x) == normalizar(c)), None
        )
        salida[c] = df[encontrada].astype(str) if encontrada else ""

    return salida.fillna(""), rutas


def seleccionar_excel():
    """Compatibilidad: en Cloud el archivo se recibe con st.file_uploader."""
    return None


def aplicar_estilo_profesional(df):
    """Estilo profesional para la vista de la tabla."""

    def estilo_fila(row):
        estilos = [""] * len(row)

        # Estado
        estado = normalizar(row.get("ESTADO DE SOLICITUD", ""))
        color_estado = {
            "ENVIADO": "#166534",
            "ENTREGADO": "#15803D",
            "POR ENVIAR": "#B45309",
            "POR REASIGNAR": "#7C3AED",
            "POR EXTRAER": "#0369A1",
            "CANCELADO": "#991B1B",
            "ANULADO": "#DC2626",
            "OBSERVADO": "#9A3412",
        }.get(estado, "#334155")

        # Prioridad
        prioridad = normalizar(row.get("PRIORIDAD", ""))
        color_prioridad = {
            "RUSH": "#B91C1C",
            "TURNO SIGUIENTE": "#B45309",
            "NORMAL": "#2563EB",
        }.get(prioridad, "#475569")

        for i, col in enumerate(row.index):
            if col == "ESTADO DE SOLICITUD":
                estilos[i] = (
                    f"background-color:{color_estado};"
                    "color:white;font-weight:700;text-align:center;"
                )
            elif col == "PRIORIDAD":
                estilos[i] = (
                    f"background-color:{color_prioridad};"
                    "color:white;font-weight:700;text-align:center;"
                )
            elif col == "CLIENTE":
                estilos[i] = "font-weight:700;color:#E2E8F0;"
            elif col == "NRO SOLICITUD - WO":
                estilos[i] = "font-weight:600;color:#93C5FD;"
            else:
                estilos[i] = "color:#CBD5E1;"
        return estilos

    return (
        df.style.apply(estilo_fila, axis=1)
        .set_properties(
            **{
                "background-color": "#0B1220",
                "border-color": "#263449",
                "font-size": "12px",
                "padding": "7px",
            }
        )
        .set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#17263D"),
                        ("color", "#F8FAFC"),
                        ("font-weight", "800"),
                        ("font-size", "11px"),
                        ("border-color", "#334155"),
                        ("text-transform", "uppercase"),
                    ],
                },
                {
                    "selector": "tbody tr:hover",
                    "props": [
                        ("background-color", "#16243A"),
                    ],
                },
            ]
        )
    )


def archivo_excel_bloqueado(ruta):
    """En Cloud no existe Microsoft Excel local; el bloqueo es por sesión."""
    with CLOUD["lock"]:
        return CLOUD.get("owner_id") is not None


def lock_path(ruta):
    return Path("/tmp/jc_control_solicitudes.edit.lock")


def write_lock_path(ruta):
    return Path("/tmp/jc_control_solicitudes.write.lock")


def leer_bloqueo_edicion(ruta):
    with CLOUD["lock"]:
        if CLOUD.get("owner_id") is None:
            return None
        return {
            "owner_id": CLOUD.get("owner_id"),
            "usuario": CLOUD.get("usuario") or "otro usuario",
            "hora_inicio": CLOUD.get("hora_inicio") or "",
            "timestamp": time.time(),
            "motivo": CLOUD.get("motivo") or "registro",
        }


def adquirir_bloqueo_edicion(ruta, motivo="registro"):
    """Reserva el dataset compartido para una sesión de Streamlit."""
    usuario = (
        st.session_state.get("usuario_nombre")
        or st.session_state.get("usuario_login")
        or "Usuario"
    ).strip()
    with CLOUD["lock"]:
        owner = CLOUD.get("owner_id")
        if owner and owner != st.session_state.session_id:
            return False, leer_bloqueo_edicion(ruta)
        CLOUD["owner_id"] = st.session_state.session_id
        CLOUD["usuario"] = usuario
        CLOUD["hora_inicio"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        CLOUD["motivo"] = motivo
        st.session_state.registro_activo = True
        return True, leer_bloqueo_edicion(ruta)


def renovar_bloqueo_edicion(ruta):
    if not st.session_state.get("registro_activo"):
        return
    with CLOUD["lock"]:
        if CLOUD.get("owner_id") == st.session_state.session_id:
            CLOUD["hora_inicio"] = CLOUD.get("hora_inicio") or datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def liberar_bloqueo_edicion(ruta):
    with CLOUD["lock"]:
        if CLOUD.get("owner_id") == st.session_state.session_id:
            CLOUD["owner_id"] = None
            CLOUD["usuario"] = None
            CLOUD["hora_inicio"] = None
            CLOUD["motivo"] = None
    st.session_state.registro_activo = False


@contextmanager
def bloqueo_guardado(ruta, espera=20):
    """Bloqueo exclusivo corto para guardar el XLSX compartido."""
    inicio = time.time()
    while True:
        with CLOUD["lock"]:
            owner = CLOUD.get("owner_id")
            if owner and owner != st.session_state.session_id:
                info = leer_bloqueo_edicion(ruta)
                usuario = info.get("usuario", "otro usuario") if info else "otro usuario"
                raise PermissionError(f"El archivo está reservado por {usuario}. Debe esperar a que termine.")
            # RLock permite que el mismo hilo/sesión continúe guardando.
            CLOUD["write_owner"] = st.session_state.session_id
            break
        if time.time() - inicio >= espera:
            raise TimeoutError("Otro usuario está guardando el Excel. Espera unos segundos.")
        time.sleep(0.2)
    try:
        yield
    finally:
        with CLOUD["lock"]:
            CLOUD.pop("write_owner", None)


def aplicar_estilo_rutas(df):
    """Colores para PROGRAMACION_RUTAS, adaptándose a las columnas del Excel."""
    if df is None or df.empty:
        return df

    regiones = {}
    paleta = [
        "#17365D",
        "#375623",
        "#7030A0",
        "#7F6000",
        "#1F4E78",
        "#385723",
        "#5B2C6F",
        "#7B3F00",
    ]

    col_region = next(
        (
            c
            for c in df.columns
            if any(x in normalizar(c) for x in ["REGIONAL", "REGION"])
        ),
        None,
    )

    def estilo_fila(row):
        estilos = [""] * len(row)

        if col_region:
            region = normalizar(row.get(col_region, ""))
            if region:
                if region not in regiones:
                    regiones[region] = paleta[len(regiones) % len(paleta)]
                color = regiones[region]
                for i in range(len(row)):
                    estilos[i] = (
                        f"background-color:{color};"
                        "color:white;border-bottom:1px solid #263449;"
                    )

        for i, col in enumerate(row.index):
            ncol = normalizar(col)
            valor = normalizar(row.get(col, ""))

            if "ESTADO" in ncol:
                if any(x in valor for x in ["ENTREGADO", "ENVIADO", "COMPLETADO"]):
                    estilos[i] = "background-color:#166534;color:white;font-weight:700;"
                elif any(x in valor for x in ["PENDIENTE", "POR ENVIAR", "POR EXTRAER"]):
                    estilos[i] = "background-color:#B45309;color:white;font-weight:700;"
                elif any(x in valor for x in ["CANCELADO", "ANULADO", "POR ETIQUETAR"]):
                    estilos[i] = "background-color:#991B1B;color:white;font-weight:700;"

            if "PRIORIDAD" in ncol and "RUSH" in valor:
                estilos[i] = "background-color:#B91C1C;color:white;font-weight:800;"

        return estilos

    return (
        df.style.apply(estilo_fila, axis=1)
        .set_properties(
            **{
                "background-color": "#0B1220",
                "color": "#E2E8F0",
                "border-color": "#263449",
                "font-size": "11px",
                "padding": "6px",
            }
        )
        .set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#17263D"),
                        ("color", "#FFFFFF"),
                        ("font-weight", "800"),
                        ("font-size", "10px"),
                        ("border-color", "#334155"),
                    ],
                },
                {
                    "selector": "tbody tr:hover",
                    "props": [("background-color", "#16243A")],
                },
            ]
        )
    )


def guardar_excel_directo(ruta, df, mensaje="Actualizar solicitudes desde Streamlit"):
    """Actualiza SOLICITUDES y publica el XLSX completo en GitHub."""
    ruta = Path(ruta)
    if ruta.suffix.lower() != ".xlsx":
        raise ValueError("El guardado requiere un archivo .xlsx.")
    with bloqueo_guardado(ruta):
        wb = load_workbook(ruta)
        if "SOLICITUDES" in wb.sheetnames:
            ws = wb["SOLICITUDES"]
            for fila in ws.iter_rows():
                for celda in fila:
                    celda.value = None
        else:
            ws = wb.create_sheet("SOLICITUDES", 0)
        for col_num, nombre in enumerate(COLS, start=1):
            ws.cell(1, col_num, nombre)
        for row_num, fila in enumerate(df[COLS].fillna("").itertuples(index=False), start=2):
            for col_num, valor in enumerate(fila, start=1):
                ws.cell(row_num, col_num, str(valor))
        # PROGRAMACION_RUTAS no se toca.
        wb.save(ruta)
        subir_excel_github(ruta, mensaje=mensaje)


def excel_bytes(df, rutas=None):
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="SOLICITUDES", index=False)
        if rutas is not None:
            rutas.to_excel(writer, sheet_name=RUTA_SHEET, index=False)
    return out.getvalue()


def badge_estado(v):
    return str(v)


# =========================================================
# LOGIN — 4 USUARIOS
# =========================================================
def obtener_usuarios_login():
    """Lee los 4 usuarios desde Streamlit Secrets.

    Configuración esperada:

    [usuarios]
    usuario1 = "clave1"
    usuario2 = "clave2"
    usuario3 = "clave3"
    usuario4 = "clave4"
    """
    try:
        datos = st.secrets.get("usuarios", {})
        return {str(k).strip().lower(): str(v) for k, v in dict(datos).items() if str(k).strip()}
    except Exception:
        return {}


def cerrar_sesion():
    """Libera el bloqueo y vuelve al login."""
    try:
        if st.session_state.get("excel_path"):
            liberar_bloqueo_edicion(st.session_state.excel_path)
    except Exception:
        pass

    st.session_state.rows = None
    st.session_state.rutas = None
    st.session_state.filename = ""
    st.session_state.excel_path = ""
    st.session_state.editing = None
    st.session_state.editing_key = ""
    st.session_state.registro_activo = False
    st.session_state.usuario_nombre = ""
    st.session_state.usuario_login = ""
    st.session_state.autenticado = False
    st.session_state.logged_out = False
    st.session_state.page = "solicitudes"
    st.rerun()


def mostrar_login():
    st.markdown(
        """
        <div style="margin:10px auto 8px auto;text-align:center;">
            <div style="font-size:28px;">🔐</div>
            <div style="font-size:22px; font-weight:700; margin:0;">Iniciar sesión</div>
            <div style="font-size:11px;opacity:.60;">Acceso al sistema de despacho</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    usuarios = obtener_usuarios_login()
    if len(usuarios) != 4:
        st.error(
            f"❌ Debes configurar exactamente 4 usuarios en Secrets. "
            f"Actualmente hay {len(usuarios)}."
        )
        return
    # FORMULARIO CENTRO Y COMPACTO
    col_izq, col_login, col_der = st.columns([1.8, 1, 1.8])
    with col_login:
        
        with st.form("login_jc_control", clear_on_submit=False):
            usuario = st.text_input(
                "👤 Usuario",
                placeholder="Ingrese su usuario",
                key="login_usuario",
            ).strip()
            password = st.text_input(
                "🔑 Contraseña",
                type="password",
                placeholder="Ingrese su contraseña",
                key="login_password",
            )
            ingresar = st.form_submit_button(
                "🔐 INGRESAR",
                use_container_width=True,
            )

    if ingresar:
        usuario_key = usuario.lower()
        if usuario_key in usuarios and password == usuarios[usuario_key]:
            st.session_state.autenticado = True
            st.session_state.usuario_login = usuario_key
            st.session_state.usuario_nombre = usuario.strip()
            st.session_state.logged_out = False
            st.rerun()
        else:
            st.error("❌ Usuario o contraseña incorrectos.")


# =========================================================
# CABECERA Y CONTROL DE ACCESO
# =========================================================
st.markdown(
    '<div class="titulo">📝 JC CONTROL DE SOLICITUDES — DESPACHO</div>',
    unsafe_allow_html=True,
)

if not st.session_state.autenticado:
    mostrar_login()
    st.stop()

# =========================================================
# IDENTIDAD DE USUARIO / EXCEL CENTRAL EN GITHUB
# =========================================================
u1, u2, u3 = st.columns([1, 3, 1])
with u1:
    st.markdown(
        f"👤 **Usuario:** `{st.session_state.usuario_nombre}`"
    )
with u2:
    st.markdown(
        f"☁️ **Excel central:** `{GITHUB_EXCEL}`"
    )
with u3:
    st.markdown('<div style="height:28px"></div>', unsafe_allow_html=True)
    salir = st.button("🚪 SALIR", use_container_width=True, key="salir_sistema")

if salir:
    cerrar_sesion()

if st.session_state.rows is None:
    try:
        if not github_configurado():
            st.error("❌ Falta GITHUB_TOKEN en los Secrets de Streamlit Cloud.")
        else:
            descargar_excel_github(CLOUD_PATH)
            st.session_state.excel_path = str(CLOUD_PATH)
            st.session_state.filename = GITHUB_EXCEL
            st.session_state.rows, st.session_state.rutas = cargar_excel_desde_ruta(CLOUD_PATH)
            st.toast("☁️ Excel central cargado desde GitHub.", icon="☁️")
    except Exception as e:
        st.error(f"❌ No se pudo cargar el Excel desde GitHub: {e}")

c1, c2, c3 = st.columns([1, 1, 4])
with c1:
    if st.button("🔄 ACTUALIZAR PAGINA", use_container_width=True):
        try:
            descargar_excel_github(CLOUD_PATH)
            st.session_state.excel_path = str(CLOUD_PATH)
            st.session_state.filename = GITHUB_EXCEL
            rows, rutas = cargar_excel_desde_ruta(CLOUD_PATH)
            st.session_state.rows = rows
            st.session_state.rutas = rutas
            st.session_state.editing = None
            st.session_state.editing_key = ""
            st.session_state.form_version += 1
            st.session_state.tabla_version += 1
            st.toast("🔄 Datos actualizados desde GitHub.", icon="🔄")
            st.rerun()
        except Exception as e:
            st.error(f"❌ No se pudo actualizar desde GitHub: {e}")
with c2:
    if st.button("🗓️ PROGRAMACIÓN DE RUTAS", use_container_width=True):
        if st.session_state.rows is not None:
            liberar_bloqueo_edicion(st.session_state.excel_path)
            st.session_state.editing = None
            st.session_state.editing_key = ""
            st.session_state.page = "rutas"
            st.rerun()
        else:
            st.warning("Primero debes cargar un archivo Excel.")

# =========================================================
# PROGRAMACIÓN DE RUTAS
# =========================================================
if st.session_state.page == "rutas":
    st.subheader("🗓️ Programación de Rutas")
    st.caption("Consulta la programación mensual de rutas.")

    if st.button("⬅ VOLVER A SOLICITUDES"):
        st.session_state.page = "solicitudes"
        st.rerun()

    # Recarga la hoja de rutas cada vez que se entra a esta vista.
    # Esto permite ver datos agregados por otro usuario sin borrar los existentes.
    if st.session_state.excel_path and Path(st.session_state.excel_path).exists():
        try:
            _, rutas_actualizadas = cargar_excel_desde_ruta(st.session_state.excel_path)
            st.session_state.rutas = rutas_actualizadas
        except Exception:
            pass

    rutas = st.session_state.rutas
    if rutas is None or rutas.empty:
        st.info(f"No hay datos en {RUTA_SHEET}. Verifica que la hoja exista.")
    else:
        col1, col2 = st.columns([1, 2])
        with col1:
            meses = {
                "TODOS": None,
                "Enero": 1,
                "Febrero": 2,
                "Marzo": 3,
                "Abril": 4,
                "Mayo": 5,
                "Junio": 6,
                "Julio": 7,
                "Agosto": 8,
                "Septiembre": 9,
                "Octubre": 10,
                "Noviembre": 11,
                "Diciembre": 12,
            }
            mes_nombre = st.selectbox("Mes", list(meses.keys()))
        with col2:
            buscar_ruta = st.text_input("Buscar", placeholder="Buscar ruta, cliente...")

        # =====================================================
        # FILTROS DE PROGRAMACIÓN DE RUTAS
        # =====================================================
        datos = rutas.copy()

        # Primero filtramos por mes usando TODAS las columnas que
        # contengan FECHA. Esto evita que el filtro falle cuando
        # la primera columna de fecha está vacía.
        mes = meses[mes_nombre]

        if mes is not None:
            columnas_fecha = [c for c in datos.columns if "FECHA" in normalizar(c)]

            if columnas_fecha:
                mascara_mes = pd.Series(False, index=datos.index)

                for columna in columnas_fecha:
                    serie = datos[columna]

                    # Intento 1: fechas normales DD/MM/YYYY, YYYY-MM-DD, etc.
                    fechas = pd.to_datetime(serie, dayfirst=True, errors="coerce")

                    # Intento 2: fechas almacenadas como número serial de Excel.
                    numeros = pd.to_numeric(serie, errors="coerce")
                    mascara_excel = fechas.isna() & numeros.notna()

                    if mascara_excel.any():
                        fechas_excel = pd.to_datetime(
                            numeros[mascara_excel],
                            unit="D",
                            origin="1899-12-30",
                            errors="coerce",
                        )
                        fechas.loc[mascara_excel] = fechas_excel

                    mascara_mes |= fechas.dt.month.eq(mes).fillna(False)

                datos = datos.loc[mascara_mes]

            else:
                # Si el Excel tiene una columna MES en lugar de una fecha,
                # también permitimos filtrar directamente por nombre.
                columna_mes = next(
                    (
                        c
                        for c in datos.columns
                        if normalizar(c) in ["MES", "MES DE RUTA", "MES RUTA"]
                    ),
                    None,
                )

                if columna_mes:
                    nombre_mes = mes_nombre.upper()

                    datos = datos[
                        datos[columna_mes]
                        .astype(str)
                        .str.strip()
                        .str.upper()
                        .eq(nombre_mes)
                    ]
                else:
                    st.warning(
                        "⚠️ No encontré una columna de fecha o MES "
                        "en PROGRAMACION_RUTAS."
                    )

        # Búsqueda automática mientras escribes.
        if buscar_ruta:
            texto_busqueda = normalizar(buscar_ruta)

            mask_busqueda = (
                datos.astype(str)
                .apply(lambda col: col.map(lambda x: texto_busqueda in normalizar(x)))
                .any(axis=1)
            )

            datos = datos.loc[mask_busqueda]

        st.caption(f"📅 {mes_nombre} · " f"{len(datos)} registros encontrados")

        st.dataframe(
            aplicar_estilo_rutas(datos),
            use_container_width=True,
            hide_index=True,
            height=520,
        )

else:
    # =====================================================
    # ARCHIVO EXCEL
    # =====================================================
    st.markdown("### 📁 Archivo Excel compartido")
    if st.session_state.rows is not None:
        st.caption(f"📄 {st.session_state.filename or 'Excel compartido'} · Datos compartidos entre las sesiones activas.")
        bloqueo = leer_bloqueo_edicion(st.session_state.excel_path)
        if bloqueo and bloqueo.get("owner_id") != st.session_state.session_id:
            st.warning(
                f"🔒 **ARCHIVO EN USO — {bloqueo.get('usuario', 'otro usuario')} está registrando o editando.** "
                "Puedes consultar los datos, pero espera antes de iniciar una nueva solicitud."
            )
    else:
        st.info("Selecciona un archivo Excel arriba para comenzar.")

    if st.session_state.rows is None:
        st.markdown(
            '<div class="footer"><strong>JC Control de Solicitudes — Almacén</strong><br>©JuanCarlosRamos - 2026 — Todos los derechos reservados</div>',
            unsafe_allow_html=True,
        )
        st.stop()

    df = st.session_state.rows

    # =====================================================
    # FORMULARIO
    # =====================================================
    st.markdown("### Registrar nueva solicitud")

    editing = st.session_state.editing

    # =====================================================
    # BLOQUEO DE EDICIÓN / NUEVO REGISTRO
    # =====================================================
    bloqueo_actual = leer_bloqueo_edicion(st.session_state.excel_path)

    if editing is None and not st.session_state.registro_activo:
        if (
            bloqueo_actual
            and bloqueo_actual.get("owner_id") != st.session_state.session_id
        ):
            usuario = bloqueo_actual.get("usuario", "otro usuario")
            equipo = bloqueo_actual.get("equipo", "")

            st.warning(
                f"🔒 **ARCHIVO EN USO — {usuario} está {'editando una solicitud' if bloqueo_actual.get('motivo') == 'edición de solicitud' else 'ingresando una solicitud'}"
                f"{' desde ' + equipo if equipo else ''}.** "
                "Espera a que termine para ingresar una nueva solicitud."
            )
        else:
            st.info(
                "🔐 Para comenzar una nueva solicitud, pulsa "
                "**INICIAR NUEVA SOLICITUD**. El archivo quedará reservado "
                "para ti mientras registras los datos."
            )

            if st.button("🔐 INICIAR NUEVA SOLICITUD", use_container_width=False):
                ok, info = adquirir_bloqueo_edicion(
                    st.session_state.excel_path,
                    motivo="nueva solicitud",
                )

                if ok:
                    st.session_state.registro_activo = True
                    st.session_state.form_version += 1
                    st.rerun()
                else:
                    usuario = (
                        info.get("usuario", "otro usuario") if info else "otro usuario"
                    )
                    equipo = info.get("equipo", "") if info else ""

                    st.warning(
                        f"🔒 **{usuario} está ingresando una solicitud.** "
                        f"{'Equipo: ' + equipo + '. ' if equipo else ''}"
                        "Espera a que termine."
                    )

    # Si estamos editando, adquirimos el bloqueo al seleccionar la fila.
    if editing is not None:
        ok, info = adquirir_bloqueo_edicion(
            st.session_state.excel_path,
            motivo="edición de solicitud",
        )

        if not ok:
            usuario = info.get("usuario", "otro usuario") if info else "otro usuario"
            equipo = info.get("equipo", "") if info else ""

            st.warning(
                f"🔒 **No puedes editar todavía. {usuario} está editando una solicitud.** "
                f"{'Equipo: ' + equipo + '. ' if equipo else ''}"
                "Espera a que termine."
            )

            st.session_state.editing = None
            st.session_state.editing_key = ""
            st.stop()

        renovar_bloqueo_edicion(st.session_state.excel_path)

        st.success(
            "🔐 **EDITANDO EN EXCLUSIVA.** Otros usuarios pueden consultar, "
            "pero deberán esperar para agregar, editar o eliminar registros."
        )

    puede_formulario = editing is not None or st.session_state.registro_activo

    if puede_formulario:
        if editing is not None and editing in df.index:
            r = df.loc[editing]
            titulo_form = "✏️ Editar solicitud"

            valores = {
                "cliente": r["CLIENTE"],
                "numero": r["NRO SOLICITUD - WO"],
                "tipo": r["TIPO DE SOLICITUD"],
                "prioridad": r["PRIORIDAD"],
                "cantidad_items": r["CANT - ITEMS"],
                "estado": r["ESTADO DE SOLICITUD"],
                "direccion": r["DIRECCIÓN"],
                "fecha": str(r["FECHA DE INGRESO"]),
            }
        else:
            titulo_form = "Registrar nueva solicitud"

            valores = {
                "cliente": CLIENTES[0],
                "numero": "",
                "tipo": TIPOS[0],
                "prioridad": PRIORIDADES[0],
                "cantidad_items": 0,
                "estado": ESTADOS[0],
                "direccion": DIRECCIONES[0],
                "fecha": str(date.today()),
            }

        with st.form(
            f"solicitud_form_{st.session_state.form_version}",
            clear_on_submit=False,
        ):
            st.markdown(f"**{titulo_form}**")

            a, b = st.columns(2)

            with a:
                cliente = st.selectbox(
                    "Cliente",
                    CLIENTES,
                    index=(
                        CLIENTES.index(valores["cliente"])
                        if valores["cliente"] in CLIENTES
                        else 0
                    ),
                )

                numero = st.text_input(
                    "Nro. Solicitud / WO",
                    value=valores["numero"],
                    placeholder="Ejemplo: SE2-26-17658",
                )

                tipo = st.selectbox(
                    "Tipo de Solicitud",
                    TIPOS,
                    index=(
                        TIPOS.index(valores["tipo"]) if valores["tipo"] in TIPOS else 0
                    ),
                )

                prioridad = st.selectbox(
                    "Prioridad",
                    PRIORIDADES,
                    index=(
                        PRIORIDADES.index(valores["prioridad"])
                        if valores["prioridad"] in PRIORIDADES
                        else 0
                    ),
                )

            with b:
                try:
                    cantidad_default = int(
                        float(str(valores.get("cantidad_items", 0)).replace(",", "."))
                    )
                except Exception:
                    cantidad_default = 0

                cantidad_items = st.number_input(
                    "Cantidad de ítems",
                    min_value=0,
                    step=1,
                    value=max(0, cantidad_default),
                    format="%d",
                )

                estado = st.selectbox(
                    "Estado de Solicitud",
                    ESTADOS,
                    index=(
                        ESTADOS.index(valores["estado"])
                        if valores["estado"] in ESTADOS
                        else 0
                    ),
                )

                direccion = st.selectbox(
                    "Dirección",
                    DIRECCIONES,
                    index=(
                        DIRECCIONES.index(valores["direccion"])
                        if valores["direccion"] in DIRECCIONES
                        else 0
                    ),
                )

                try:
                    fecha_default = pd.to_datetime(
                        valores["fecha"],
                        dayfirst=True,
                        errors="coerce",
                    ).date()

                    if pd.isna(fecha_default):
                        fecha_default = date.today()

                except Exception:
                    fecha_default = date.today()

                fecha = st.date_input(
                    "Fecha de ingreso",
                    value=fecha_default,
                )

            x, y = st.columns([4, 1])

            with x:
                guardar = st.form_submit_button(
                    (
                        "💾 ACTUALIZAR REGISTRO"
                        if editing is not None
                        else "📥 AGREGAR REGISTRO"
                    ),
                    use_container_width=True,
                )

            with y:
                cancelar = st.form_submit_button(
                    "✕ CANCELAR",
                    use_container_width=True,
                )
    else:
        guardar = False
        cancelar = False

    if cancelar:
        liberar_bloqueo_edicion(st.session_state.excel_path)
        st.session_state.editing = None
        st.session_state.editing_key = ""
        st.session_state.form_version += 1
        st.session_state.tabla_version += 1
        st.rerun()

    if guardar:
        numero_norm = normalizar(numero)

        if not numero_norm or not all(
            [
                cliente != CLIENTES[0],
                tipo != TIPOS[0],
                prioridad != PRIORIDADES[0],
                int(cantidad_items) >= 0,
                estado != ESTADOS[0],
                direccion != DIRECCIONES[0],
            ]
        ):
            st.error("Completa todos los campos antes de guardar.")
        else:
            try:
                # IMPORTANTE:
                # Antes de guardar se vuelve a leer el Excel real.
                # Así no se sobrescriben datos que otro usuario haya agregado.
                latest_df, latest_rutas = cargar_excel_desde_ruta(
                    st.session_state.excel_path
                )

                nuevo = {
                    "CLIENTE": cliente,
                    "NRO SOLICITUD - WO": numero.strip(),
                    "TIPO DE SOLICITUD": tipo,
                    "PRIORIDAD": prioridad,
                    "CANT - ITEMS": str(int(cantidad_items)),
                    "ESTADO DE SOLICITUD": estado,
                    "DIRECCIÓN": direccion,
                    "FECHA DE INGRESO": fecha.strftime("%d/%m/%Y"),
                }

                if editing is not None:
                    # Buscar por Nro. Solicitud / WO, no por posición.
                    # Esto evita editar otra fila si otro usuario agregó/eliminó registros.
                    clave_edicion = st.session_state.editing_key or normalizar(
                        valores["numero"]
                    )

                    coincidencias = [
                        i
                        for i, v in latest_df["NRO SOLICITUD - WO"].items()
                        if normalizar(v) == clave_edicion
                    ]

                    if not coincidencias:
                        st.warning(
                            "⚠️ Este registro cambió en el Excel mientras lo editabas. "
                            "Se recargaron los datos. Vuelve a seleccionarlo para editar."
                        )
                        st.session_state.rows = latest_df
                        st.session_state.rutas = latest_rutas
                        liberar_bloqueo_edicion(st.session_state.excel_path)
                        st.session_state.editing = None
                        st.session_state.editing_key = ""
                        st.session_state.form_version += 1
                        st.session_state.tabla_version += 1
                        st.rerun()

                    indice_real = coincidencias[0]

                    # Si el nuevo WO pertenece a otra solicitud, validar duplicado.
                    repetido = any(
                        i != indice_real and normalizar(v) == numero_norm
                        for i, v in latest_df["NRO SOLICITUD - WO"].items()
                    )

                    if repetido:
                        st.error(
                            "Ya existe otra solicitud con ese Nro. Solicitud / WO."
                        )
                    else:
                        latest_df.loc[indice_real, COLS] = [nuevo[c] for c in COLS]
                        guardar_excel_directo(
                            st.session_state.excel_path,
                            latest_df,
                            mensaje=f"Guardar solicitud {numero_norm} desde Streamlit",
                        )

                        # Recargar después de guardar para mostrar exactamente
                        # lo que quedó en el archivo.
                        rows_final, rutas_final = cargar_excel_desde_ruta(
                            st.session_state.excel_path
                        )
                        st.session_state.rows = rows_final
                        st.session_state.rutas = rutas_final
                        liberar_bloqueo_edicion(st.session_state.excel_path)
                        st.session_state.editing = None
                        st.session_state.editing_key = ""
                        st.session_state.form_version += 1
                        st.session_state.tabla_version += 1
                        st.success(
                            "✅ Registro actualizado. Se conservaron todos los demás datos."
                        )
                        st.rerun()

                else:
                    repetido = any(
                        normalizar(v) == numero_norm
                        for v in latest_df["NRO SOLICITUD - WO"]
                    )

                    if repetido:
                        st.error("Ya existe una solicitud con ese Nro. Solicitud / WO.")
                    else:
                        latest_df = pd.concat(
                            [latest_df, pd.DataFrame([nuevo])],
                            ignore_index=True,
                        )

                        guardar_excel_directo(
                            st.session_state.excel_path,
                            latest_df,
                            mensaje=f"Guardar solicitud {numero_norm} desde Streamlit",
                        )

                        # Recargar también PROGRAMACION_RUTAS.
                        # No se elimina ni se reemplaza esa hoja.
                        rows_final, rutas_final = cargar_excel_desde_ruta(
                            st.session_state.excel_path
                        )
                        st.session_state.rows = rows_final
                        st.session_state.rutas = rutas_final
                        liberar_bloqueo_edicion(st.session_state.excel_path)
                        st.session_state.form_version += 1
                        st.session_state.tabla_version += 1
                        st.success(
                            "✅ Nueva solicitud agregada. No se eliminó ningún dato del Excel."
                        )
                        st.rerun()

            except PermissionError as e:
                st.warning(
                    "🔒 El archivo está siendo utilizado por otro usuario o "
                    "está abierto en Microsoft Excel. Espera a que termine y vuelve a guardar."
                )
            except TimeoutError:
                st.warning(
                    "⏳ Otro usuario está guardando el Excel en este momento. "
                    "Espera unos segundos y vuelve a intentarlo."
                )
            except Exception as e:
                liberar_bloqueo_edicion(st.session_state.excel_path)
                st.error(f"❌ No se pudo guardar el Excel: {e}")

    # st.markdown("### 💾 Guardar / descargar")

    d1, d2 = st.columns(2)

    with d1:
        if st.button("💾 GUARDAR / ACTUALIZAR EXCEL", use_container_width=True):
            try:
                guardar_excel_directo(
                    st.session_state.excel_path,
                    st.session_state.rows,
                    mensaje="Sincronizar solicitudes desde Streamlit",
                )
                rows_final, rutas_final = cargar_excel_desde_ruta(st.session_state.excel_path)
                st.session_state.rows = rows_final
                st.session_state.rutas = rutas_final
                liberar_bloqueo_edicion(st.session_state.excel_path)
                st.session_state.editing = None
                st.session_state.editing_key = ""
                st.session_state.form_version += 1
                st.session_state.tabla_version += 1
                st.toast("☁️ Excel actualizado en GitHub.", icon="☁️")
                st.rerun()
            except PermissionError as e:
                st.warning(f"🔒 {e}")
            except TimeoutError as e:
                st.warning(f"⏳ {e}")
                try:
                    descargar_excel_github(CLOUD_PATH)
                    st.session_state.rows, st.session_state.rutas = cargar_excel_desde_ruta(CLOUD_PATH)
                except Exception:
                    pass
            except Exception as e:
                st.error(f"❌ No se pudo actualizar el Excel en GitHub: {e}")

    # =====================================================
    # RESULTADOS / FILTRO
    # =====================================================
    st.markdown("### Resultados")

    # -----------------------------------------------------
    # BÚSQUEDA + FILTROS: ESTADO + MES/DÍA
    # -----------------------------------------------------
    buscar_col, estado_col, fecha_col = st.columns([3, 1, 1])

    with buscar_col:
        buscar = st.text_input(
            "Buscar solicitud...",
            placeholder="Escribe aquí y el filtro se actualizará automáticamente...",
            key="buscar_solicitud_live",
        )

    with estado_col:
        estados_filtro = ["TODOS"] + ESTADOS[1:]
        filtro_estado = st.selectbox(
            "Estado de solicitud",
            estados_filtro,
            key="filtro_estado_solicitud",
        )

    with fecha_col:
        filtro_fecha = st.date_input(
            "📅 Filtrar fecha",
            value=None,
            key="filtro_fecha_solicitud",
        )

    vista = df.copy()

    # Búsqueda automática mientras se escribe.
    if buscar:
        mask = (
            vista.astype(str)
            .apply(
                lambda col: col.map(
                    lambda x: normalizar(x).find(normalizar(buscar)) >= 0
                )
            )
            .any(axis=1)
        )
        vista = vista[mask]

    # Filtro desplegable por estado.
    if filtro_estado != "TODOS":
        vista = vista[
            vista["ESTADO DE SOLICITUD"].map(normalizar).eq(normalizar(filtro_estado))
        ]

    # -----------------------------------------------------
    # FILTRO POR FECHA — UN SOLO CALENDARIO
    # -----------------------------------------------------
    if filtro_fecha:
        fechas_vista = pd.to_datetime(
            vista["FECHA DE INGRESO"],
            dayfirst=True,
            errors="coerce",
        )
        vista = vista[fechas_vista.dt.date.eq(filtro_fecha)]

    estado_caption = f" · Estado: {filtro_estado}" if filtro_estado != "TODOS" else ""
    fecha_caption = f" · 📅 {filtro_fecha.strftime('%d/%m/%Y')}" if filtro_fecha else ""

    st.caption(
        f"{len(vista)} de {len(df)} registros · 🔎 filtrado automático al escribir"
        f"{estado_caption}{fecha_caption}"
    )

    # -----------------------------------------------------
    # TABLA CON ACCIONES EN LA MISMA FILA
    # -----------------------------------------------------
    if len(vista):
        tabla = vista[COLS].copy()

        # Guardamos el índice real del DataFrame para que editar/eliminar
        # siga funcionando correctamente aunque exista un filtro.
        tabla["_INDICE_REAL_"] = vista.index

        # Columnas de acción que aparecen dentro de la misma tabla.
        # Indicadores visuales profesionales para estado y prioridad.
        # Se conservan los valores originales para guardar en Excel.
        tabla["ESTADO DE SOLICITUD"] = tabla["ESTADO DE SOLICITUD"].apply(
            lambda x: (
                "🟢 " + str(x)
                if normalizar(x) in ["ENTREGADO", "ENVIADO"]
                else (
                    "🟡 " + str(x)
                    if normalizar(x) in ["POR ENVIAR", "TURNO SIGUIENTE"]
                    else (
                        "🔵 " + str(x)
                        if normalizar(x) in ["POR EXTRAER"]
                        else (
                            "🟣 " + str(x)
                            if normalizar(x) in ["POR REASIGNAR"]
                            else (
                                "🔴 " + str(x)
                                if normalizar(x) in ["CANCELADO", "ANULADO"]
                                else "⚪ " + str(x)
                            )
                        )
                    )
                )
            )
        )
        tabla["PRIORIDAD"] = tabla["PRIORIDAD"].apply(
            lambda x: (
                "🔴 " + str(x)
                if normalizar(x) == "RUSH"
                else (
                    "🟠 " + str(x)
                    if normalizar(x) == "TURNO SIGUIENTE"
                    else "🔵 " + str(x) if normalizar(x) == "NORMAL" else "⚪ " + str(x)
                )
            )
        )

        tabla["✏️ EDITAR"] = False
        tabla["🗑️ ELIMINAR"] = False

        st.markdown(
            """
            <div style="
                display:flex;
                gap:18px;
                flex-wrap:wrap;
                margin:6px 0 10px;
                padding:8px 12px;
                background:#0B1220;
                border:1px solid #263449;
                border-radius:6px;
                color:#CBD5E1;
                font-size:11px;">
                <b style="color:#E2E8F0;">ESTADO:</b>
                🟢 Enviado/Entregado
                🟡 Por Enviar
                🔵 Por extraer
                🟣 Por Reasignar
                🔴 Anulado
                <b style="color:#E2E8F0;">PRIORIDAD:</b>
                🔴 RUSH
                🟠 Turno siguiente
            </div>
            """,
            unsafe_allow_html=True,
        )

        resultado_tabla = st.data_editor(
            tabla,
            use_container_width=True,
            hide_index=True,
            height=450,
            key=f"tabla_solicitudes_{st.session_state.tabla_version}",
            disabled=COLS + ["_INDICE_REAL_"],
            column_config={
                "_INDICE_REAL_": None,
                "CANT - ITEMS": st.column_config.NumberColumn(
                    "ÍTEMS",
                    help="Cantidad de ítems de la solicitud",
                    min_value=0,
                    step=1,
                    format="%d",
                ),
                "✏️ EDITAR": st.column_config.CheckboxColumn(
                    "✏️ EDITAR",
                    help="Marca esta casilla para editar este registro",
                    default=False,
                ),
                "🗑️ ELIMINAR": st.column_config.CheckboxColumn(
                    "🗑️ ELIMINAR",
                    help="Marca esta casilla para eliminar este registro",
                    default=False,
                ),
            },
        )

        # Detectar si el usuario marcó EDITAR.
        # _INDICE_REAL_ conserva el índice REAL de df, incluso cuando hay filtros.
        editar_marcados = resultado_tabla.loc[resultado_tabla["✏️ EDITAR"] == True]

        if not editar_marcados.empty:
            indice_real = editar_marcados.iloc[0]["_INDICE_REAL_"]

            ok, info = adquirir_bloqueo_edicion(
                st.session_state.excel_path,
                motivo="edición de solicitud",
            )

            if not ok:
                usuario = (
                    info.get("usuario", "otro usuario") if info else "otro usuario"
                )
                equipo = info.get("equipo", "") if info else ""

                st.warning(
                    f"🔒 **{usuario} ya está registrando/editando una solicitud.** "
                    f"{'Equipo: ' + equipo + '. ' if equipo else ''}"
                    "No se modificó nada. Espera a que termine."
                )
            else:
                st.session_state.editing = indice_real
                st.session_state.editing_key = normalizar(
                    df.loc[indice_real, "NRO SOLICITUD - WO"]
                )
                st.session_state.form_version += 1
                st.session_state.tabla_version += 1
                st.rerun()

        # Detectar si el usuario marcó ELIMINAR.
        eliminar_marcados = resultado_tabla[resultado_tabla["🗑️ ELIMINAR"] == True]

        if len(eliminar_marcados):
            try:
                ok, info = adquirir_bloqueo_edicion(
                    st.session_state.excel_path,
                    motivo="eliminación de solicitud",
                )

                if not ok:
                    usuario = (
                        info.get("usuario", "otro usuario") if info else "otro usuario"
                    )
                    equipo = info.get("equipo", "") if info else ""

                    st.warning(
                        f"🔒 **{usuario} está utilizando el archivo.** "
                        f"{'Equipo: ' + equipo + '. ' if equipo else ''}"
                        "No se eliminó ningún registro."
                    )
                    st.stop()

                # Usamos el Nro. Solicitud / WO como identificador estable.
                wos_eliminar = {
                    normalizar(x)
                    for x in vista.loc[
                        eliminar_marcados["_INDICE_REAL_"].tolist(),
                        "NRO SOLICITUD - WO",
                    ].tolist()
                }

                latest_df, latest_rutas = cargar_excel_desde_ruta(
                    st.session_state.excel_path
                )

                mask_eliminar = latest_df["NRO SOLICITUD - WO"].map(
                    lambda x: normalizar(x) in wos_eliminar
                )
                latest_df = latest_df.loc[~mask_eliminar].reset_index(drop=True)

                guardar_excel_directo(
                    st.session_state.excel_path,
                    latest_df,
                    mensaje="Eliminar registros desde Streamlit",
                )

                rows_final, rutas_final = cargar_excel_desde_ruta(
                    st.session_state.excel_path
                )
                st.session_state.rows = rows_final
                st.session_state.rutas = rutas_final
                liberar_bloqueo_edicion(st.session_state.excel_path)

                if st.session_state.editing_key in wos_eliminar:
                    st.session_state.editing = None
                    st.session_state.editing_key = ""

                st.success(
                    "✅ Registro(s) eliminado(s). Se conservaron las demás hojas y rutas."
                )

                st.session_state.form_version += 1
                st.session_state.tabla_version += 1
                st.rerun()

            except PermissionError:
                st.warning(
                    "🔒 El archivo está siendo utilizado por otro usuario "
                    "o está abierto en Microsoft Excel."
                )
            except TimeoutError:
                st.warning(
                    "⏳ Otro usuario está guardando el Excel. Espera unos segundos."
                )
            except Exception as e:
                st.error(f"❌ No se pudo guardar el Excel: {e}")

    else:
        st.info("No hay registros que coincidan con la búsqueda.")

    # =====================================================
    # DESCARGAS
    # =====================================================

    # d1, d2 = st.columns(2)

    with d2:
        filtrados_bytes = excel_bytes(vista[COLS], None)
        nombre_filtrado = f"SOLICITUDES_FILTRADAS_{date.today().isoformat()}.xlsx"
        st.download_button(
            "📥 DESCARGAR FILTRADOS",
            data=filtrados_bytes,
            file_name=nombre_filtrado,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.markdown(
    '<div class="footer"><strong>JC Control de Solicitudes — Almacén</strong><br>©JuanCarlosRamos - 2026 — Todos los derechos reservados</div>',
    unsafe_allow_html=True,
)
